import os.path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from captura_preco import captura_preco, armazena_link
from datetime import date
from winotify import Notification, audio

# Guardando a data do dia em que o script é executado e formatando para o formato brasileiro
data = date.today()
dataFormatada = data.strftime('%d/%m/%Y')
local_planilha= r'C:\Users\GuiMo\Downloads\monitoramento'
arq_planilha = os.path.join(local_planilha,'price_monitor.xlsx')

ultima_linha = sheet.max_row
proxima_linha = ultima_linha + 1
celula_a = sheet[f'A{proxima_linha}']
celula_b = sheet[f'B{proxima_linha}']
ultimo_valor_a = sheet[f'A{ultima_linha}']




def cria_planilha():
        #Verifica se o arquivo existe dentro da pasta do script. Caso não exista, ele cria a pasta, e o arquivo.
    if not os.path.exists(arq_planilha):
        global sheet
        global workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'VALOR DO PROUTO:'
        sheet['B1'] = 'DATA DO VALOR:'
        workbook.save(arq_planilha)
    else:
        workbook = load_workbook(arq_planilha)
        sheet = workbook.active

def verifica_ultima_l():

    # Captura a última linha preenchida com texto na planilha e verifica a próxima linha vazia para inserir os dados

    # Verifica o valor da última célula A com preço

    # Verifica se a célula atual está vazia, e anota o preço e data na próxima linha
    if celula_a.value is None:
        celula_a.value = captura_preco()

    if celula_b.value is None:
        celula_b.value = dataFormatada
        workbook.save(arq_planilha)

def notifys():
        # Configura notificações

    notify_up = Notification(app_id="Monitorador de Preço", title="O PREÇO DO PRODUTO SUBIU",
                                 msg=f'O preço atual é: {celula_a.value} reais.',
                                 duration="short",
                                 icon="icons_noti/high_price.jpg")

    notify_up.set_audio(audio.Reminder, loop=False)

    notify_up.add_actions(label="Link para o produto",
                              launch=armazena_link())

    notify_down = Notification(app_id="Monitorador de Preço", title="O PREÇO DO PRODUTO CAIU",
                                   msg=f'O preço atual é: {celula_a.value} reais.',
                                   icon="icons_noti/low_price.jpg")

    notify_down.set_audio(audio.Reminder, loop=False)

    notify_down.add_actions(label="Link para o produto",
                                launch=armazena_link())
    if celula_a.value and ultimo_valor_a.value:  # Verifica se os valores não são None
        if celula_a.value > ultimo_valor_a.value:
                notify_up.show()
        elif celula_a.value < ultimo_valor_a.value:
                notify_down.show()

    else:
        print("A ultima célula verificada está em branco, ou vazia.")

        preco_historico = sheet['A2']
        new_preco_historico = float(preco_historico.value)

    if new_preco_historico > celula_a.value:
        # Verificação e criação de um preço histórico para monitoramento do menor preço já atingido dentro da planilha
        preco_historico = celula_a.value
        notify_record = Notification(app_id="Monitorador de Preço", title="MENOR PREÇO HISTÓRICO",
                                             msg=f'O PREÇO ATUAL É: {celula_a.value} REAIS!!',
                                             icon="icons_noti/record_lowPrice.gif")

        notify_record.set_audio(audio.Reminder, loop=False)
        notify_record.add_actions(label="Link para o produto",
                                          launch=armazena_link())
        notify_record.show()


def anota_preco():
    cria_planilha()
    verifica_ultima_l()
    notifys()


anota_preco()
